# /// script
# requires-python = ">=3.11"
# dependencies = ["polars>=1.0", "httpx>=0.27", "openpyxl>=3.1", "tenacity>=8.0"]
# ///
"""Scrape FMCSA acute/critical (Serious) Violations from investigations.

These come from the per-carrier SMS XLSX download (proxy-free), the ONLY source
for investigation Serious Violations — they are in no bulk file. Feeds ISS
Group 6 + the "Serious Violation sets the BASIC percentile to 100" rule.

Endpoint: https://ai.fmcsa.dot.gov/SMS/Carrier/{DOT}/Download.aspx?BASIC=0&FileType=XLSX
Sheet:    "Acute-Critical Violations" → Investigation Type | Date | Violation | Description | BASIC

Scope (likely-investigated carriers; investigations are rare so this is a tight
superset of carriers that could have a Serious Violation):
    FAST Act high-risk  ∪  compliance-reviewed in last 12mo  ∪  has enforcement case

Output: data/fmcsa_scrape/serious_violations_<SMS_DATA_TAG>.parquet
    one row per (DOT, violation): dot_number, investigation_type,
    investigation_date, violation_code, description, basic, scraped_at
Plus per-DOT status in serious_violations_status_*.parquet (ok/no_data/error)
for resume + coverage tracking.
"""
from __future__ import annotations

import asyncio
import io
import os
import sys
import time
from datetime import datetime
from pathlib import Path

import httpx
import openpyxl
import polars as pl
from tenacity import retry, stop_after_attempt, wait_exponential, retry_if_exception_type

REPO = Path(__file__).resolve().parents[2]
# Repo-relative, not a hardcoded checkout. These defaults pointed at the
# san-antonio workspace, so a standalone `uv run` silently read from — or
# wrote into — a different clone. Fine under build_all (which supplies the
# env); wrong every other way.
PARQUET = Path(os.environ.get("FMCSA_PARQUET", REPO / "data" / "carrier_aggregates.parquet"))
CENSUS = Path(os.environ.get("FMCSA_COMPANY_CENSUS", "/__unset__run-via-build_all.py-or-set-the-env-var__/Company_Census_File.csv"))
OUT_DIR = Path(os.environ.get("FMCSA_SCRAPE_DIR", REPO / "data" / "fmcsa_scrape"))
# Vintage tag for the output filename. Was hardcoded, so an August scrape
# wrote into a file named ...20260514 and the name silently lied about the
# data. Reads SMS_DATA_TAG like the rest of the pipeline. NOTE: changing
# the tag resets the resume set, since progress is keyed on the status
# file of the same tag.
TAG = os.environ.get("SMS_DATA_TAG", "20260514")
URL = "https://ai.fmcsa.dot.gov/SMS/Carrier/{dot}/Download.aspx?BASIC=0&FileType=XLSX"
UA = "augment-carrier-audit-research/0.1 (+research@goaugment.com; polite scraper)"
# Env-tunable: the default of 8 is ~5x slower than the ZenRows proxy sustains.
# The PU-history scrape measured 1.0 rps at 8 workers vs 5.1 rps at 48, with
# the error rate FALLING (0.4% -> 0.1%), so 8 is leaving throughput on the
# table rather than protecting us from the WAF.
CONCURRENCY = int(os.environ.get("SCRAPE_CONCURRENCY", "8"))
CHECKPOINT_EVERY = 300

# Wall-clock budget; 0/unset = unlimited. See scrape_pu_history for the why: a
# hosted CI job is capped at 6h and the full refresh no longer fits, so the
# scrape spans runs. Past the budget we cancel outstanding work, checkpoint, and
# exit EXIT_PARTIAL — every finished DOT is durable and the resume logic above
# (status file: ok / no_data are done) picks up the rest next run.
#
# NOTE the asymmetry with scrape_pu_history: that scraper has a global RPS
# limiter, so its CONCURRENCY only governs whether it reaches a fixed ceiling.
# This one has NO rate cap — here CONCURRENCY *is* the request rate, so raising
# SCRAPE_CONCURRENCY makes us genuinely more aggressive. Left at 8 deliberately.
#
# SCRAPE_DEADLINE_EPOCH is absolute (unix seconds) and is what CI sets, so the
# budget is SHARED with the other scrape step rather than granted twice.
SCRAPE_BUDGET_MIN = float(os.environ.get("SCRAPE_BUDGET_MIN", "0") or 0)
SCRAPE_DEADLINE_EPOCH = float(os.environ.get("SCRAPE_DEADLINE_EPOCH", "0") or 0)
EXIT_PARTIAL = 75


def _deadline_epoch() -> float | None:
    """Earliest of the absolute deadline and the relative budget, or None."""
    candidates = []
    if SCRAPE_DEADLINE_EPOCH > 0:
        candidates.append(SCRAPE_DEADLINE_EPOCH)
    if SCRAPE_BUDGET_MIN > 0:
        candidates.append(time.time() + SCRAPE_BUDGET_MIN * 60)
    return min(candidates) if candidates else None


def candidate_dots() -> list[int]:
    df = pl.read_parquet(PARQUET)
    hr = set(df.filter(pl.col("fast_act_high_risk"))["DOT_NUMBER"].to_list())
    enf = set(df.filter(pl.col("enforcement_cases_count") > 0)["DOT_NUMBER"].to_list())
    cen = (pl.scan_csv(CENSUS, ignore_errors=True, infer_schema_length=3000)
           .select("DOT_NUMBER", pl.col("REVIEW_DATE").cast(pl.Int64, strict=False).alias("rd"))
           .collect())
    reviewed = set(cen.filter(pl.col("rd") >= 20250501)["DOT_NUMBER"].to_list())
    return sorted(hr | enf | reviewed)


def parse_acute_critical(dot: int, data: bytes) -> tuple[str, list[dict]]:
    """Return (status, rows). status: 'ok' (had rows), 'no_data' (empty), error."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        return f"error:parse:{type(e).__name__}", []
    if "Acute-Critical Violations" not in wb.sheetnames:
        return "no_sheet", []
    ws = wb["Acute-Critical Violations"]
    rows = []
    it = ws.iter_rows(values_only=True)
    header = next(it, None)  # Investigation Type | Date | Violation | Description | BASIC
    for r in it:
        if not r or all(c is None for c in r):
            continue
        rows.append({
            "dot_number": dot,
            "investigation_type": str(r[0]) if r[0] is not None else None,
            "investigation_date": str(r[1]) if len(r) > 1 and r[1] is not None else None,
            "violation_code": str(r[2]) if len(r) > 2 and r[2] is not None else None,
            "description": str(r[3])[:200] if len(r) > 3 and r[3] is not None else None,
            "basic": str(r[4]) if len(r) > 4 and r[4] is not None else None,
        })
    return ("ok" if rows else "no_data"), rows


class Fetcher:
    def __init__(self):
        # The Download.aspx endpoint tolerates only a few hundred UNPROXIED
        # requests before FMCSA's WAF blocks the IP (observed: ~400 then a
        # 403 storm). Route through ZenRows when SCRAPE_PROXY_URL is set:
        #   export SCRAPE_PROXY_URL="<key>:mode=auto@api.zenrows.com:8001"
        # Nothing ever exported SCRAPE_PROXY_URL — build_all.py doesn't set it,
        # so this scraper ran UNPROXIED every month while scrape_pu_history.py
        # (which falls back to ZENROWS_API_KEY on its own) ran proxied. That
        # asymmetry is the likely reason this scrape failed 714/5,006 = 14.3% of
        # DOTs in Aug 2026 against the CI scraper's 0.5%: the WAF was blocking
        # us. Mirror scrape_pu_history and derive the proxy from the credential
        # we already have.
        proxy = os.environ.get("SCRAPE_PROXY_URL")
        if not proxy:
            zr_key = os.environ.get("ZENROWS_API_KEY")
            if zr_key:
                params = os.environ.get("ZENROWS_PROXY_PARAMS", "premium_proxy=true")
                proxy = f"http://{zr_key}:{params}@api.zenrows.com:8001"
        kwargs = dict(headers={"User-Agent": UA}, follow_redirects=True,
                      timeout=90.0, limits=httpx.Limits(max_connections=CONCURRENCY))
        if proxy:
            url = proxy if proxy.startswith("http") else f"http://{proxy}"
            kwargs["proxy"] = url
            kwargs["verify"] = False  # ZenRows MITMs TLS
            print(f"[proxy] routing via {url.split('@')[-1]}", flush=True)
        elif os.environ.get("ALLOW_UNPROXIED_SCRAPE") == "1":
            print("[proxy] none — UNPROXIED by explicit opt-in (WAF blocks ~400 reqs)", flush=True)
        else:
            # Refuse rather than warn. An unproxied run doesn't just fail, it
            # burns the IP for later runs, and the failure is silent: WAF
            # rejections get recorded as ordinary per-DOT errors.
            raise SystemExit(
                "fetch_serious_violations: no proxy. FMCSA's WAF blocks direct IPs after ~400 "
                "requests and the rejections are written as per-DOT errors, so the scrape looks "
                "like it merely had a bad day. Set ZENROWS_API_KEY (or SCRAPE_PROXY_URL), or "
                "ALLOW_UNPROXIED_SCRAPE=1 to override."
            )
        self.client = httpx.AsyncClient(**kwargs)
        self.sem = asyncio.Semaphore(CONCURRENCY)

    # Only 429 used to raise, so tenacity never saw anything else and every other
    # non-200 became a PERMANENT verdict on the first try. ZenRows answers 422
    # when its upstream fetch fails — a transient proxy condition, not "this DOT
    # has no page". The Aug 2026 run wrote 714 of 5,006 DOTs (14.3%) off as
    # error:http422; retried later, they returned 200 with full pages. Those
    # carriers were simply never checked for serious violations.
    RETRYABLE_STATUS = {408, 422, 425, 429, 500, 502, 503, 504}

    # reraise=True so an exhausted retry surfaces the ORIGINAL httpx error
    # ("http422") instead of tenacity's RetryError, which would erase the status
    # code and record every exhausted DOT as an indistinguishable "RetryError".
    @retry(stop=stop_after_attempt(3),
           wait=wait_exponential(multiplier=1, min=2, max=20),
           retry=retry_if_exception_type((httpx.HTTPError, httpx.TimeoutException)),
           reraise=True)
    async def _get(self, dot: int) -> httpx.Response:
        r = await self.client.get(URL.format(dot=dot))
        if r.status_code in self.RETRYABLE_STATUS:
            if r.status_code == 429:
                await asyncio.sleep(30)
            raise httpx.HTTPError(f"http{r.status_code}")
        return r

    async def fetch(self, dot: int) -> tuple[int, str, list[dict]]:
        async with self.sem:
            now = time.strftime("%Y-%m-%dT%H:%M:%S")
            try:
                r = await self._get(dot)
            except Exception as e:
                # Keep the status a SHORT, GROUPABLE code. It used to embed the
                # exception text (and in the CI scraper, the full URL), so every
                # failure was its own distinct value and a group_by returned
                # hundreds of singleton rows instead of one "error: N" —
                # which is why a 14% failure rate went unnoticed for months.
                msg = str(e)
                code = msg if msg.startswith("http") else type(e).__name__
                return dot, f"error:{code}", []
            if r.status_code != 200:
                return dot, f"error:http{r.status_code}", []
            status, rows = parse_acute_critical(dot, r.content)
            for row in rows:
                row["scraped_at"] = now
            return dot, status, rows

    async def close(self):
        await self.client.aclose()


def write_checkpoint(all_rows: list[dict], status: list[dict]):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    schema = {"dot_number": pl.Int64, "investigation_type": pl.Utf8, "investigation_date": pl.Utf8,
              "violation_code": pl.Utf8, "description": pl.Utf8, "basic": pl.Utf8, "scraped_at": pl.Utf8}
    pl.DataFrame(all_rows, schema=schema).write_parquet(OUT_DIR / f"serious_violations_{TAG}.parquet", compression="zstd")
    pl.DataFrame(status, schema={"dot_number": pl.Int64, "scrape_status": pl.Utf8, "n_sv": pl.Int64}) \
        .write_parquet(OUT_DIR / f"serious_violations_status_{TAG}.parquet", compression="zstd")


async def main():
    dots = candidate_dots()
    # Resume: skip DOTs already in the status file.
    status_path = OUT_DIR / f"serious_violations_status_{TAG}.parquet"
    done = set()
    all_rows, status = [], []
    if status_path.exists():
        prev = pl.read_parquet(status_path)
        # Only ok / no_data count as done — RETRY errored DOTs (e.g. WAF 403s
        # from an unproxied run). Keep their non-error rows + statuses.
        keep = prev.filter(pl.col("scrape_status").is_in(["ok", "no_data"]))
        done = set(keep["dot_number"].to_list())
        status = keep.to_dicts()
        sv_path = OUT_DIR / f"serious_violations_{TAG}.parquet"
        if sv_path.exists():
            all_rows = pl.read_parquet(sv_path).filter(
                pl.col("dot_number").is_in(list(done))
            ).to_dicts()
    todo = [d for d in dots if d not in done]
    print(f"candidates: {len(dots):,}  already done: {len(done):,}  todo: {len(todo):,}", flush=True)

    f = Fetcher()
    t0 = time.monotonic(); ok = nod = err = 0
    deadline = _deadline_epoch()  # epoch seconds, or None for unlimited
    deferred = 0
    try:
        tasks = [asyncio.create_task(f.fetch(d)) for d in todo]
        for i, coro in enumerate(asyncio.as_completed(tasks), 1):
            dot, st, rows = await coro
            all_rows.extend(rows)
            status.append({"dot_number": dot, "scrape_status": st, "n_sv": len(rows)})
            if st == "ok": ok += 1
            elif st == "no_data": nod += 1
            else: err += 1
            if i % CHECKPOINT_EVERY == 0:
                write_checkpoint(all_rows, status)
                rps = i / (time.monotonic() - t0)
                print(f"  {i}/{len(todo)}  ok={ok} nodata={nod} err={err}  {rps:.1f}/s  "
                      f"ETA {(len(todo)-i)/rps/60:.0f}m", flush=True)
            if deadline is not None and time.time() >= deadline:
                # Cancel rather than drain: unlike the PU scrape these are all
                # real in-flight requests, and we want the budget to mean the
                # wall clock, not "however long the tail takes to finish".
                pending = [t for t in tasks if not t.done()]
                deferred = len(pending)
                for t in pending:
                    t.cancel()
                # Reap them so the loop's exit doesn't surface CancelledError.
                await asyncio.gather(*tasks, return_exceptions=True)
                print(f"\nPARTIAL: time budget elapsed with "
                      f"{deferred:,} DOTs unscraped.", flush=True)
                break
    finally:
        await f.close()
        write_checkpoint(all_rows, status)
    print(f"\nDONE. carriers with Serious Violations: {ok:,}  "
          f"total SV rows: {len(all_rows):,}  no_data: {nod:,}  errors: {err:,}")
    if all_rows:
        sv = pl.DataFrame(all_rows)
        print("\nSerious Violations by BASIC:")
        print(sv.group_by("basic").len().sort("len", descending=True))

    if deferred:
        print(f"  Re-run to resume — {deferred:,} DOTs remain.", flush=True)
        return EXIT_PARTIAL
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
